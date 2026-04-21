# app/api/v1/users.py
import io
import json
import secrets
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_password_hash, create_setup_token
from app.api.v1.auth import get_current_user, _user_material_scopes
from app.models.user import User
from app.crud.user import get_users, get_user_by_id, get_user_by_username, create_user, update_user_role, approve_user, reject_user, delete_user, update_user_permissions
from app.crud.operation_log import log_operation
from app.schemas.auth import UserListItem, UserUpdateRoleRequest, UserApproveRequest, UserUpdatePermissionsRequest, AdminCreateUserRequest, AdminCreateUserResponse, VALID_ROLES
from app.core.permissions import get_user_module_permissions, VALID_PERM_LEVELS, LEGACY_MODULE_IDS, MANAGED_MODULES, has_level

router = APIRouter()

ROLE_LABELS = {
    "admin": "超级管理员",
    "electrical_requisition_clerk": "电气领用员",
    "mechanical_requisition_clerk": "机械领用员",
    "electrical_admin": "电气管理员",
    "mechanical_admin": "机械管理员",
    "general_staff": "通用人员",
    "requisition_clerk": "备件领用员",  # 兼容旧数据
}

# 分配以下角色前，必须先为用户配置对应模块权限（先设模块权限，再分配角色）
ROLE_MODULE_REQUIREMENT: dict[str, str] = {
    "electrical_admin": "electrical",
    "electrical_requisition_clerk": "electrical",
    "mechanical_admin": "mechanical",
    "mechanical_requisition_clerk": "mechanical",
}


def _role_label(role: str) -> str:
    return ROLE_LABELS.get(role, role or "—")


ROLE_ALIAS_MAP: dict[str, str] = {
    "admin": "admin",
    "超级管理员": "admin",
    "electrical_requisition_clerk": "electrical_requisition_clerk",
    "电气领用员": "electrical_requisition_clerk",
    "requisition_clerk": "electrical_requisition_clerk",
    "机械领用员": "mechanical_requisition_clerk",
    "mechanical_requisition_clerk": "mechanical_requisition_clerk",
    "electrical_admin": "electrical_admin",
    "电气管理员": "electrical_admin",
    "mechanical_admin": "mechanical_admin",
    "机械管理员": "mechanical_admin",
    "general_staff": "general_staff",
    "通用人员": "general_staff",
}


def require_can_manage_users(current_user: User = Depends(get_current_user)) -> User:
    """仅超级管理员、电气管理员、机械管理员可访问用户管理"""
    if current_user.role not in ("admin", "electrical_admin", "mechanical_admin"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="仅管理员可操作用户管理")
    return current_user


def _allowed_roles_for_approve_and_update(current_user: User) -> tuple[str, ...]:
    """当前用户可审批/修改为的角色：超级管理员任意（含 general_staff）；电气管理员仅电气领用员；机械管理员仅机械领用员"""
    if current_user.username == "admin":
        return VALID_ROLES
    if current_user.role == "electrical_admin":
        return ("electrical_requisition_clerk",)
    if current_user.role == "mechanical_admin":
        return ("mechanical_requisition_clerk",)
    return ()


def _allowed_roles_for_create(current_user: User) -> tuple[str, ...]:
    """当前用户新建用户时可指定的角色：超级管理员任意；电气/机械管理员可建本模块领用员或通用人员。"""
    if current_user.username == "admin":
        return VALID_ROLES
    if current_user.role == "electrical_admin":
        return ("electrical_requisition_clerk", "general_staff")
    if current_user.role == "mechanical_admin":
        return ("mechanical_requisition_clerk", "general_staff")
    return ()


def _user_to_item(u: User) -> UserListItem:
    return UserListItem(
        id=u.id,
        username=u.username,
        real_name=getattr(u, "real_name", None),
        role=u.role,
        material_scopes=_user_material_scopes(u),
        status=getattr(u, "status", "approved") or "approved",
        wechat_userid=getattr(u, "wechat_userid", None),
        wechat_name=getattr(u, "wechat_name", None),
        permissions=get_user_module_permissions(u),
        created_at=u.created_at,
    )


def _filter_users_by_viewer(users: list[User], viewer: User) -> list[User]:
    """
    按查看者身份过滤用户列表：
    - admin：可见所有用户
    - 电气管理员：可见所有具有电气模块权限（角色或 permissions JSON）的用户，以及待审批用户
    - 机械管理员：可见所有具有机械模块权限（角色或 permissions JSON）的用户，以及待审批用户
    """
    if viewer.role == "admin":
        return users
    # 非 admin 时，admin 用户对其他人不可见
    visible = [u for u in users if u.role != "admin"]
    if viewer.role == "electrical_admin":
        visible = [
            u for u in visible
            if "electrical" in _user_material_scopes(u)
            or getattr(u, "status", "approved") == "pending"
        ]
    elif viewer.role == "mechanical_admin":
        visible = [
            u for u in visible
            if "mechanical" in _user_material_scopes(u)
            or getattr(u, "status", "approved") == "pending"
        ]
    return visible


@router.get("/users", response_model=list[UserListItem], tags=["用户管理"])
async def list_users(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_can_manage_users),
):
    """获取用户列表。admin 可见全部；电气/机械管理员仅可见本模块管理员与领用员；admin 对非 admin 不可见。"""
    users = get_users(db)
    visible = _filter_users_by_viewer(users, current_user)
    return [_user_to_item(u) for u in visible]


@router.post("/users", response_model=AdminCreateUserResponse, status_code=status.HTTP_201_CREATED, tags=["用户管理"])
async def create_user_endpoint(
    body: AdminCreateUserRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_can_manage_users),
):
    """
    管理员新建用户。授权分级：
    - 超级管理员：可建任意角色、任意模块权限。
    - 电气管理员：仅可建电气领用员或通用人员；通用人员仅能配置电气模块且最高 editor。
    - 机械管理员：仅可建机械领用员或通用人员；通用人员仅能配置机械模块且最高 editor。
    新建后返回设置密码链接用 token，请将链接发给用户自助设置密码（7 天内有效）。
    """
    allowed_roles = _allowed_roles_for_create(current_user)
    if body.role not in allowed_roles:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="您只能新建以下角色：" + "、".join(_role_label(r) for r in allowed_roles),
        )
    if get_user_by_username(db, body.username.strip()):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="用户名已存在")

    # 通用人员时校验 permissions：仅允许当前管理员可配置的模块与级别
    perms = body.permissions or {}
    if body.role == "general_staff" and perms:
        is_super = current_user.username == "admin"
        managed = MANAGED_MODULES.get(current_user.role)
        for mod_id, level in perms.items():
            if not level:
                continue
            if level not in VALID_PERM_LEVELS:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"权限级别 '{level}' 不合法，可选：editor / viewer",
                )
            if not is_super:
                if not managed or mod_id not in managed["modules"]:
                    raise HTTPException(
                        status_code=status.HTTP_403_FORBIDDEN,
                        detail=f"您无权为该用户配置 {mod_id} 模块的权限",
                    )
                if not has_level(managed["max_level"], level):
                    raise HTTPException(
                        status_code=status.HTTP_403_FORBIDDEN,
                        detail=f"您最高只能授予 {managed['max_level']} 级别，无法授予 {level}",
                    )

    placeholder_hash = get_password_hash(secrets.token_urlsafe(32))
    u = create_user(
        db,
        username=body.username.strip(),
        password_hash=placeholder_hash,
        role=body.role,
        status="approved",
        real_name=(body.real_name or "").strip() or None,
    )
    if body.role == "general_staff" and perms:
        update_user_permissions(db, u.id, {k: v for k, v in perms.items() if v})

    setup_token = create_setup_token(u.id)
    try:
        summary = f"新建用户 {u.username}" + (f"（{body.real_name}）" if body.real_name else "") + f"，角色：{_role_label(body.role)}"
        log_operation(
            db=db,
            user=current_user,
            module="user",
            action="create",
            entity_type="user",
            entity_id=u.id,
            summary=summary,
        )
        db.commit()
    except Exception:
        db.rollback()
    return AdminCreateUserResponse(
        username=u.username,
        real_name=getattr(u, "real_name", None),
        token=setup_token,
    )


@router.post("/users/batch", tags=["用户管理"])
async def batch_create_users_endpoint(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_can_manage_users),
):
    """
    批量新增用户（Excel）：字段顺序为「账号、姓名、角色」。
    - 角色支持中文标签或英文值（如：电气领用员 / electrical_requisition_clerk）。
    - 当前管理员仍受角色授权限制。
    """
    filename = (file.filename or "").strip()
    if not filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="缺少上传文件名")
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="仅支持 .xlsx 文件，请先另存为 xlsx 后再导入")

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="服务端缺少 openpyxl 依赖，请先安装后再导入")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="上传文件为空")

    try:
        wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel 解析失败，请检查文件是否损坏或格式不正确")

    ws = wb.worksheets[0] if wb.worksheets else None
    if ws is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="未找到有效工作表")

    allowed_roles = set(_allowed_roles_for_create(current_user))
    seen_usernames: set[str] = set()
    created_items: list[dict] = []
    errors: list[dict] = []

    def _norm(v) -> str:
        return str(v or "").strip()

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel 无数据")

    header = [_norm(v).lower() for v in (rows[0] or ())]
    username_idx = 0
    real_name_idx = 1
    role_idx = 2
    header_map = {
        "账号": "username", "用户名": "username", "username": "username", "user_name": "username",
        "姓名": "real_name", "真实姓名": "real_name", "real_name": "real_name", "name": "real_name",
        "角色": "role", "role": "role",
    }
    if header:
        mapped: dict[str, int] = {}
        for idx, h in enumerate(header):
            key = header_map.get(h)
            if key and key not in mapped:
                mapped[key] = idx
        if "username" in mapped and "role" in mapped:
            username_idx = mapped["username"]
            role_idx = mapped["role"]
            real_name_idx = mapped.get("real_name", real_name_idx)
            data_rows = rows[1:]
        else:
            data_rows = rows
    else:
        data_rows = rows

    for i, row in enumerate(data_rows, start=2 if data_rows is rows[1:] else 1):
        vals = list(row or ())
        username = _norm(vals[username_idx] if len(vals) > username_idx else "")
        real_name = _norm(vals[real_name_idx] if len(vals) > real_name_idx else "")
        role_raw = _norm(vals[role_idx] if len(vals) > role_idx else "")

        if not username and not role_raw and not real_name:
            continue
        if not username:
            errors.append({"row": i, "username": "", "error": "账号为空"})
            continue
        if len(username) > 64:
            errors.append({"row": i, "username": username, "error": "账号长度不能超过 64"})
            continue
        if username in seen_usernames:
            errors.append({"row": i, "username": username, "error": "文件内账号重复"})
            continue
        seen_usernames.add(username)

        role_key = ROLE_ALIAS_MAP.get(role_raw, "")
        if not role_key:
            errors.append({"row": i, "username": username, "error": f"角色不合法：{role_raw or '空'}"})
            continue
        if role_key not in allowed_roles:
            errors.append({"row": i, "username": username, "error": f"无权创建角色：{_role_label(role_key)}"})
            continue
        if get_user_by_username(db, username):
            errors.append({"row": i, "username": username, "error": "账号已存在"})
            continue

        try:
            placeholder_hash = get_password_hash(secrets.token_urlsafe(32))
            u = create_user(
                db,
                username=username,
                password_hash=placeholder_hash,
                role=role_key,
                status="approved",
                real_name=real_name or None,
            )
            setup_token = create_setup_token(u.id)
            created_items.append({
                "username": u.username,
                "real_name": getattr(u, "real_name", None),
                "token": setup_token,
            })
        except Exception as e:
            db.rollback()
            errors.append({"row": i, "username": username, "error": f"创建失败：{str(e)}"})

    created = len(created_items)
    failed = len(errors)

    try:
        summary = f"批量新增用户：成功 {created} 个，失败 {failed} 个"
        log_operation(
            db=db,
            user=current_user,
            module="user",
            action="batch_create",
            entity_type="user",
            entity_id=0,
            summary=summary,
        )
        db.commit()
    except Exception:
        db.rollback()

    return {
        "created": created,
        "failed": failed,
        "items": created_items,
        "errors": errors,
    }


def _viewer_can_see_user(viewer: User, target: User) -> bool:
    """当前查看者是否允许看到/操作该目标用户（与列表过滤一致）。"""
    if viewer.role == "admin":
        return True
    if target.role == "admin":
        return False
    if getattr(target, "status", "approved") == "pending":
        return viewer.role in ("electrical_admin", "mechanical_admin")
    if viewer.role == "electrical_admin":
        return "electrical" in _user_material_scopes(target)
    if viewer.role == "mechanical_admin":
        return "mechanical" in _user_material_scopes(target)
    return False


def _viewer_can_delete_user(viewer: User, target: User) -> bool:
    """
    当前用户是否可删除目标用户。
    - admin：可删非 admin 且非自己的用户（含 general_staff）
    - 电气管理员：仅可删电气领用员（不可删 general_staff）
    - 机械管理员：仅可删机械领用员
    """
    if not _viewer_can_see_user(viewer, target):
        return False
    if viewer.id == target.id:
        return False
    if target.username == "admin":
        return False
    if viewer.role == "admin":
        return True
    if viewer.role == "electrical_admin":
        return target.role == "electrical_requisition_clerk"
    if viewer.role == "mechanical_admin":
        return target.role == "mechanical_requisition_clerk"
    return False


@router.patch("/users/{user_id}", response_model=UserListItem, tags=["用户管理"])
async def update_role(
    user_id: int,
    body: UserUpdateRoleRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_can_manage_users),
):
    """更新用户角色。超级管理员可设任意角色；电气管理员只能设为电气领用员；机械管理员只能设为机械领用员。"""
    if user_id == current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="不能修改自己的权限",
        )
    u = get_user_by_id(db, user_id)
    if not u:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")
    if not _viewer_can_see_user(current_user, u):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")
    # 电气/机械管理员可见 general_staff 用户但不可修改其角色（只能配置模块权限）
    if u.role == "general_staff" and current_user.username != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="通用人员的角色只能由超级管理员修改，模块权限请使用「模块权限」按钮配置")
    if body.role not in VALID_ROLES and body.role != "requisition_clerk":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="角色须为：电气领用员、机械领用员、电气管理员、机械管理员、通用人员或超级管理员")
    allowed = _allowed_roles_for_approve_and_update(current_user)
    new_role = body.role if body.role != "requisition_clerk" else "electrical_requisition_clerk"
    if new_role not in allowed:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="您只能将用户设为：" + "、".join(_role_label(r) for r in allowed),
        )
    # 模块权限先行：分配模块角色前，用户须已具备对应模块权限
    required_module = ROLE_MODULE_REQUIREMENT.get(new_role)
    if required_module and required_module not in _user_material_scopes(u):
        module_label = "电气" if required_module == "electrical" else "机械"
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"请先为该用户配置「{module_label}备件」模块权限（只读或可编辑），再分配此角色",
        )
    u = update_user_role(db, user_id, new_role)
    if not u:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")
    try:
        name_part = getattr(u, "wechat_name", None) or u.username
        summary = f"将用户 {u.username}（{name_part}）角色修改为 {_role_label(body.role)}"
        log_operation(
            db=db,
            user=current_user,
            module="user",
            action="update_role",
            entity_type="user",
            entity_id=user_id,
            summary=summary,
        )
        db.commit()
    except Exception:
        db.rollback()
    return _user_to_item(u)


@router.post("/users/{user_id}/approve", response_model=UserListItem, tags=["用户管理"])
async def approve(
    user_id: int,
    body: UserApproveRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_can_manage_users),
):
    """审批通过：超级管理员可设任意角色（含通用人员）；电气管理员只能通过为电气领用员；机械管理员只能通过为机械领用员。
    审批为通用人员时，可在 body.permissions 中同步设置初始模块权限。
    """
    u = get_user_by_id(db, user_id)
    if not u:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在或非待审批状态")
    if not _viewer_can_see_user(current_user, u):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")
    if body.role not in VALID_ROLES and body.role != "requisition_clerk":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="角色须为：电气领用员、机械领用员、电气管理员、机械管理员、通用人员或超级管理员")
    allowed = _allowed_roles_for_approve_and_update(current_user)
    approve_role = body.role
    if body.role == "requisition_clerk":
        approve_role = "electrical_requisition_clerk"
    if approve_role not in allowed:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="您只能将新用户审批为：" + "、".join(_role_label(r) for r in allowed),
        )
    # general_staff 审批时校验初始 permissions（不能含旧版模块 key，权限值合法）
    initial_permissions = body.permissions or {}
    if approve_role == "general_staff" and initial_permissions:
        for mod_id, level in initial_permissions.items():
            if mod_id in LEGACY_MODULE_IDS:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                                    detail=f"旧版模块 {mod_id} 的权限由角色控制")
            if level not in VALID_PERM_LEVELS:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                                    detail=f"权限级别 '{level}' 不合法，可选：editor / viewer")
    # 模块权限先行：审批为模块角色前，须先在 initial_permissions 中包含对应模块权限
    required_module = ROLE_MODULE_REQUIREMENT.get(approve_role)
    if required_module and required_module not in (initial_permissions or {}):
        # 如果 initial_permissions 未提供对应模块权限，检查用户自注册信息中是否已有
        if required_module not in _user_material_scopes(u):
            module_label = "电气" if required_module == "electrical" else "机械"
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"请先为该用户配置「{module_label}备件」模块权限（只读或可编辑），再审批为此角色",
            )
    u = approve_user(db, user_id, approve_role)
    if not u:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在或非待审批状态")
    # 审批为 general_staff 时同步写入初始模块权限
    if approve_role == "general_staff" and initial_permissions:
        u = update_user_permissions(db, u.id, initial_permissions) or u
    try:
        name_part = getattr(u, "wechat_name", None) or u.username
        perm_note = f"，初始模块权限：{json.dumps(initial_permissions, ensure_ascii=False)}" if initial_permissions else ""
        summary = f"审批通过用户 {u.username}（{name_part}），角色设为 {_role_label(approve_role)}{perm_note}"
        log_operation(
            db=db,
            user=current_user,
            module="user",
            action="approve",
            entity_type="user",
            entity_id=user_id,
            summary=summary,
        )
        db.commit()
    except Exception:
        db.rollback()
    return _user_to_item(u)


@router.post("/users/{user_id}/reject", tags=["用户管理"])
async def reject(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_can_manage_users),
):
    """审批拒绝：删除待审批用户"""
    target = get_user_by_id(db, user_id)
    if not target:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在或非待审批状态")
    if not _viewer_can_see_user(current_user, target):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")
    name_part = f"{target.username}（{getattr(target, 'wechat_name', None) or target.username}）"
    if not reject_user(db, user_id):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在或非待审批状态")
    try:
        summary = f"拒绝用户 {name_part} 的注册申请"
        log_operation(
            db=db,
            user=current_user,
            module="user",
            action="reject",
            entity_type="user",
            entity_id=user_id,
            summary=summary,
        )
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "已拒绝"}


@router.patch("/users/{user_id}/permissions", response_model=UserListItem, tags=["用户管理"])
async def update_permissions_endpoint(
    user_id: int,
    body: UserUpdatePermissionsRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    配置用户模块权限。
    - 超级管理员：可设置任意模块、任意级别（admin/editor/viewer），全量替换。
    - 电气/机械管理员：只能配置 general_staff 用户的本模块权限（最高 editor），与现有权限合并。
    """
    is_super_admin = current_user.username == "admin"
    managed = MANAGED_MODULES.get(current_user.role)  # None if not a module admin

    if not is_super_admin and not managed:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="仅管理员可配置模块权限")
    if user_id == current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="不能修改自己的模块权限")

    u = get_user_by_id(db, user_id)
    if not u:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")
    if not _viewer_can_see_user(current_user, u):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")

    # 模块管理员不可配置其他管理员级别用户的权限（超级管理员除外）
    if not is_super_admin and u.role in ("electrical_admin", "mechanical_admin", "admin"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无法配置其他管理员的模块权限",
        )

    # 校验权限字段合法性
    for mod_id, level in body.permissions.items():
        if not level:
            continue  # 空字符串 = 删除该模块权限，跳过合法性校验
        if level not in VALID_PERM_LEVELS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"权限级别 '{level}' 不合法，可选：editor / viewer",
            )
        if not is_super_admin:
            if mod_id not in managed["modules"]:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail=f"您无权配置 {mod_id} 模块的权限",
                )
            if not has_level(managed["max_level"], level):
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail=f"您最高只能授予 {managed['max_level']} 级别权限，无法授予 {level}",
                )

    # 计算最终权限
    if is_super_admin:
        # 超级管理员：全量替换（过滤空值）
        final_permissions = {k: v for k, v in body.permissions.items() if v}
    else:
        # 模块管理员：合并——只更新自己管理的模块键，其余模块键保持不变
        existing = get_user_module_permissions(u)
        final_permissions = dict(existing)
        for mod_id in managed["modules"]:
            if mod_id in body.permissions:
                if body.permissions[mod_id]:
                    final_permissions[mod_id] = body.permissions[mod_id]
                else:
                    final_permissions.pop(mod_id, None)

    u = update_user_permissions(db, user_id, final_permissions)
    if not u:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")
    try:
        summary = f"更新用户 {u.username} 的模块权限：{json.dumps(final_permissions, ensure_ascii=False)}"
        log_operation(db=db, user=current_user, module="user", action="update_permissions",
                      entity_type="user", entity_id=user_id, summary=summary)
        db.commit()
    except Exception:
        db.rollback()
    return _user_to_item(u)


@router.delete("/users/{user_id}", status_code=status.HTTP_200_OK, tags=["用户管理"])
async def delete_user_endpoint(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_can_manage_users),
):
    """删除用户。超级管理员可删任意用户（除自己与 admin）；电气管理员仅可删电气领用员；机械管理员仅可删机械领用员。电气与机械隔离。"""
    u = get_user_by_id(db, user_id)
    if not u:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")
    if not _viewer_can_delete_user(current_user, u):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="您没有权限删除该用户",
        )
    name_part = getattr(u, "wechat_name", None) or u.username
    if not delete_user(db, user_id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="不能删除该用户")
    try:
        summary = f"删除用户 {u.username}（{name_part}）"
        log_operation(
            db=db,
            user=current_user,
            module="user",
            action="delete",
            entity_type="user",
            entity_id=user_id,
            summary=summary,
        )
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "已删除"}
